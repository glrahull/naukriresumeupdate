import openpyxl

def dataGenerator():
    wk = openpyxl.load_workbook("C:/Users/rahul/Downloads/login.xlsx")
    sh = wk['Sheet1']
    r = sh.max_row
    data_list = []

    for i in range(1, r+1):
        username = sh.cell(i, 1).value
        password = sh.cell(i, 2).value
        data_list.append((username, password))

    print(data_list)  # Optional: Print the data list for verification
    return data_list
